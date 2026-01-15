import base64
import os
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# =========================
# Page config
# =========================
st.set_page_config(
    page_title="Република България — BGGovAI (DEMO)",
    layout="wide",
)


# =========================
# Optional: OpenAI (real-time AI)
# Works with openai>=1.x, and falls back gracefully if missing.
# =========================
def ask_ai(system: str, context: str) -> str:
    """
    Calls OpenAI if OPENAI_API_KEY is set in Streamlit Secrets.
    Falls back to a safe local message if unavailable.
    """
    api_key = None
    try:
        api_key = st.secrets.get("OPENAI_API_KEY", None)
    except Exception:
        api_key = None

    if not api_key:
        return (
            "⚠️ Няма зададен OPENAI_API_KEY в Streamlit Secrets.\n\n"
            "Демото работи и без AI, но за *real-time* анализ добави ключ:\n"
            "Manage app → Settings → Secrets → OPENAI_API_KEY = \"...\""
        )

    model = None
    try:
        model = st.secrets.get("OPENAI_MODEL", "gpt-4o-mini")
    except Exception:
        model = "gpt-4o-mini"

    # Try OpenAI v1.x client
    try:
        from openai import OpenAI  # type: ignore
        client = OpenAI(api_key=api_key)
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": context},
            ],
            temperature=0.2,
        )
        return resp.choices[0].message.content
    except Exception:
        # Fallback to legacy openai (pre-1.0)
        try:
            import openai  # type: ignore

            openai.api_key = api_key
            resp = openai.ChatCompletion.create(
                model=model,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": context},
                ],
                temperature=0.2,
            )
            return resp["choices"][0]["message"]["content"]
        except Exception as e:
            return (
                "❌ AI повикването не мина.\n\n"
                "Най-честите причини:\n"
                "• Липсва `openai` в requirements.txt\n"
                "• Грешен/невалиден ключ\n"
                "• Моделът в OPENAI_MODEL не е достъпен\n\n"
                f"Технически детайл: {e}"
            )


# =========================
# Assets: crest (demo)
# Avoid FileNotFoundError by embedding a fallback SVG.
# =========================
ASSETS_DIR = Path(__file__).parent / "assets"

DEMO_CREST_SVG = """\
<svg xmlns="http://www.w3.org/2000/svg" width="128" height="128" viewBox="0 0 128 128">
  <defs>
    <linearGradient id="g" x1="0" x2="1">
      <stop offset="0" stop-color="#00966E"/>
      <stop offset="1" stop-color="#D62612"/>
    </linearGradient>
  </defs>
  <rect x="8" y="8" width="112" height="112" rx="18" fill="#ffffff" stroke="#111827" stroke-width="2"/>
  <path d="M64 24c18 0 34 10 34 26v18c0 20-16 36-34 36S30 88 30 68V50c0-16 16-26 34-26z"
        fill="url(#g)" stroke="#111827" stroke-width="2"/>
  <path d="M64 40l8 16 18 2-13 12 3 18-16-9-16 9 3-18-13-12 18-2z"
        fill="#ffffff" opacity="0.9"/>
  <text x="64" y="118" font-family="Arial, sans-serif" font-size="12" text-anchor="middle" fill="#111827">
    DEMO
  </text>
</svg>
"""

def b64_bytes(data: bytes) -> str:
    return base64.b64encode(data).decode("utf-8")

def load_crest_b64() -> str:
    crest_path = ASSETS_DIR / "crest_demo.svg"
    if crest_path.exists():
        try:
            return b64_bytes(crest_path.read_bytes())
        except Exception:
            pass
    return b64_bytes(DEMO_CREST_SVG.encode("utf-8"))

CREST_B64 = load_crest_b64()


# =========================
# UI Header (official-style)
# =========================
st.markdown(
    f"""
    <style>
      .gov-header {{
        border-radius: 14px;
        overflow: hidden;
        box-shadow: 0 2px 10px rgba(0,0,0,0.08);
        border: 1px solid rgba(0,0,0,0.06);
        margin-bottom: 14px;
        background: #ffffff;
      }}
      .flag {{
        height: 10px;
        background: linear-gradient(
            to bottom,
            #ffffff 0%, #ffffff 33%,
            #00966E 33%, #00966E 66%,
            #D62612 66%, #D62612 100%
        );
      }}
      .gov-top {{
        display: flex;
        gap: 14px;
        align-items: center;
        padding: 14px 16px;
      }}
      .crest {{
        width: 54px;
        height: 54px;
        flex: 0 0 54px;
      }}
      .gov-title {{
        line-height: 1.15;
      }}
      .gov-title h1 {{
        margin: 0;
        font-size: 18px;
        font-weight: 800;
        letter-spacing: 0.2px;
      }}
      .gov-title p {{
        margin: 4px 0 0 0;
        font-size: 13px;
        color: rgba(0,0,0,0.65);
      }}
      .disclaimer {{
        border-radius: 12px;
        padding: 10px 12px;
        background: rgba(214,38,18,0.06);
        border: 1px solid rgba(214,38,18,0.20);
        font-size: 13px;
        margin-bottom: 10px;
      }}
      .chip {{
        display:inline-block;
        padding: 4px 10px;
        border-radius: 999px;
        background: rgba(0,150,110,0.08);
        border: 1px solid rgba(0,150,110,0.22);
        font-size: 12px;
        margin-right: 6px;
        margin-bottom: 6px;
      }}
      .muted {{
        color: rgba(0,0,0,0.60);
        font-size: 12px;
      }}
      .box {{
        border: 1px solid rgba(0,0,0,0.08);
        border-radius: 14px;
        padding: 12px 12px;
        background: #fff;
      }}
    </style>

    <div class="gov-header">
      <div class="flag"></div>
      <div class="gov-top">
        <img class="crest" src="data:image/svg+xml;base64,{CREST_B64}" />
        <div class="gov-title">
          <h1>Република България — BGGovAI</h1>
          <p>ИИ съветник за публични политики (демонстрационна версия)</p>
        </div>
      </div>
    </div>

    <div class="disclaimer">
      <b>Внимание:</b> Това е <b>демо прототип</b>. Не е официален държавен портал и не представлява правен/финансов съвет.
      Използваното лого е <b>стилизиран DEMO символ</b>.
    </div>
    """,
    unsafe_allow_html=True,
)


# =========================
# Supported topics (demo)
# =========================
SUPPORTED = [
    "ДДС 9% за ресторанти (въздействие върху бюджета)",
    "Пенсии +10% (въздействие върху разходите)",
    "Инвестиции (Capex+образование+здраве — сценарий)",
    "Общ фискален преглед (дефицит/дълг/AIC)",
    "Закон за българското гражданство (рамка за правен анализ)",
    "Смяна на МОЛ на ЕООД (административни стъпки и документи)",
]

st.markdown("### Поддържани теми (демо валидирани)")
st.markdown(" ".join([f'<span class="chip">{s}</span>' for s in SUPPORTED]), unsafe_allow_html=True)
st.caption("Пиши свободно — системата ще разпознае темата и ще извади анализ.")


# =========================
# Inputs: question + optional Excel
# =========================
st.markdown("### Въпрос към системата")
q = st.text_area(
    "Въведи въпрос (за фискални въпроси прикачи Excel бюджета):",
    height=90,
    placeholder="Пример: Какво става ако върнем ДДС 9% за ресторанти? Какъв е ефектът върху дефицита и целите?",
)

uploaded = st.file_uploader("Качи Excel бюджет (.xlsx)", type=["xlsx"])


# =========================
# Intent classifier
# =========================
def classify(text: str) -> str:
    t = (text or "").strip().lower()

    if any(k in t for k in ["мол", "управител", "еоод", "търговски регист", "а4", "вписване", "агенция по вписванията"]):
        return "ADMIN_MOL"

    if any(k in t for k in ["гражданств", "закон за българското гражданство", "натурализ", "изменени", "проект", "чл.", "ал.", "параграф", "§"]):
        return "LEGAL_CITIZENSHIP"

    # VAT restaurants
    if "ддс" in t and any(k in t for k in ["ресторан", "кетър", "хран", "9%","9 %","девет"]):
        return "FISCAL_VAT_REST"

    if "пенс" in t and any(k in t for k in ["10", "процент", "%", "+10"]):
        return "FISCAL_PENSIONS"

    if any(k in t for k in ["инвест", "капекс", "инфраструкт", "образован", "здравеопаз", "capex"]):
        return "FISCAL_INVEST"

    if any(k in t for k in ["дефиц", "дълг", "бюджет", "бвп", "aic", "догон", "маастрихт"]):
        return "FISCAL_BASE"

    # If unclear, still try: legal/admin keywords first; else default to GENERAL
    return "GENERAL"


# =========================
# Admin & legal answers (demo)
# =========================
def answer_admin_mol():
    st.subheader("Администрация: Смяна на МОЛ (управител) на ЕООД — DEMO чеклист")
    st.markdown(
        """
**Къде:** Търговски регистър (Агенция по вписванията)  
**Заявление:** обичайно **А4** (промени по обстоятелства)

**Документи (типично):**
- Решение на едноличния собственик за освобождаване/назначаване на управител
- Съгласие и образец от подпис (спесимен) на новия управител *(често с нотариална заверка — зависи от практиката/случая)*
- Декларации по ТЗ/ЗТРРЮЛНЦ (според конкретиката и заявителя)
- При електронно подаване: КЕП (квалифициран електронен подпис)

**Стъпки:**
1) Подготвяш решение + декларации + спесимен  
2) Подаване в ТР (електронно е по-евтино)  
3) След вписване: уведомяваш банка/контрагенти, актуализираш договори/пълномощни при нужда
"""
    )
    st.caption("Бележка: демо ориентир. Реалният пакет документи зависи от конкретния казус и изискванията за заверки.")


def answer_legal_citizenship():
    st.subheader("Право: Закон за българското гражданство — DEMO рамка за анализ")
    st.markdown(
        """
**Как да оцениш предложение за промяна (структура за анализ):**
1) **Точен обхват**: кои текстове (чл./ал./§) се променят и как  
2) **Конституционност / съответствие**: с Конституция, международни договори, принципи на правовата държава  
3) **Процедури и изпълнимост**: срокове, доказване на условия, натоварване на администрацията, контрол  
4) **Рискове**: неясни дефиниции, широко усмотрение, обжалвания, конфликт на норми, празноти в преходни разпоредби  
5) **Мерки за минимизиране**: ясни дефиниции, преходни правила, подзаконови актове, ИТ/регистрови промени, стандарти за доказване

Ако искаш **точен правен анализ**, копирай тук текста на предложението (или конкретните членове) и ще маркирам:
- какво реално се променя
- потенциални противоречия/рискове
- практически ефект върху процедурата
"""
    )


# =========================
# Excel parsing helpers
# Expect sheets: Inputs, Revenues, Expenditures
# =========================
def table_to_df(rows, total_keyword="TOTAL"):
    header = None
    body = []
    for r in rows:
        if r and len(r) >= 2 and r[0] == "Category" and r[1] == "Amount (bn BGN)":
            header = list(r[:3])
            continue
        if header and r and r[0]:
            body.append(list(r[:3]))
    df = pd.DataFrame(body, columns=header or ["Category", "Amount (bn BGN)", "Notes"])
    df = df[~df["Category"].astype(str).str.contains(total_keyword, na=False)]
    df["Amount (bn BGN)"] = pd.to_numeric(df["Amount (bn BGN)"], errors="coerce").fillna(0.0)
    return df


def parse_inputs(rows):
    vals = {}
    for r in rows:
        if not r or not r[0]:
            continue
        vals[str(r[0]).strip()] = r[1]

    def getf(k, default=None):
        v = vals.get(k, default)
        try:
            return float(v)
        except Exception:
            return default

    return {
        "gdp": getf("GDP (bn BGN)", None),
        "debt": getf("Debt stock (bn BGN)", None),
        "aic_bg": getf("AIC (EU=100) - Bulgaria", 70.0),
        "aic_eu": getf("AIC (EU=100) - EU average", 100.0),
    }


def traffic(deficit_pct, debt_pct, goal_def=0.03, goal_debt=0.60):
    def light(val, green_th, yellow_th):
        if val is None:
            return "⚪️"
        if val <= green_th:
            return "🟩"
        if val <= yellow_th:
            return "🟨"
        return "🟥"

    f = light(abs(deficit_pct) if deficit_pct is not None else None, goal_def, goal_def * 1.5)
    d = light(debt_pct, goal_debt, goal_debt + 0.10)
    return f, d


# =========================
# Run button
# =========================
do = st.button("Отговори", use_container_width=True)

if not do:
    st.stop()

intent = classify(q)


# =========================
# General: if unclear, still answer using AI (no Excel required)
# =========================
GOALS_TEXT = """\
Цели (демо):
- Дефицит ≤ 3% от БВП (Маастрихт)
- Дълг ≤ 60% от БВП (Маастрихт)
- Максимално бързо догонване по AIC (ЕС=100)
- Без вдигане на данъци (като политическо ограничение)
"""


# =========================
# Fiscal block requires Excel
# =========================
if intent.startswith("FISCAL"):
    if not uploaded:
        st.warning("За финансовите въпроси първо качи Excel бюджета (.xlsx).")
        st.stop()

    wb = load_workbook(filename=BytesIO(uploaded.getvalue()), data_only=True)
    need = {"Inputs", "Revenues", "Expenditures"}
    if not need.issubset(set(wb.sheetnames)):
        st.error("Липсват нужни листове. Нужни са: Inputs, Revenues, Expenditures.")
        st.stop()

    inp = parse_inputs(list(wb["Inputs"].values))
    gdp = inp["gdp"]
    debt = inp["debt"]
    aic_bg = inp["aic_bg"]
    aic_eu = inp["aic_eu"]

    rev_df = table_to_df(list(wb["Revenues"].values), total_keyword="TOTAL")
    exp_df = table_to_df(list(wb["Expenditures"].values), total_keyword="TOTAL")

    goal_def = 0.03
    goal_debt = 0.60

    # Apply demo scenario deltas (intentionally simple / fictive)
    note = "DEMO: общ фискален преглед (без промяна)."

    if intent == "FISCAL_VAT_REST":
        # Example: reduce VAT revenue by fictive 0.6 bn
        rev_df.loc[rev_df["Category"] == "VAT (total)", "Amount (bn BGN)"] -= 0.6
        note = "DEMO сценарий: ДДС 9% за ресторанти → -0.6 млрд. лв. от общ ДДС (условно)."

    elif intent == "FISCAL_PENSIONS":
        exp_df.loc[exp_df["Category"] == "Pensions", "Amount (bn BGN)"] *= 1.10
        note = "DEMO сценарий: +10% пенсии → увеличение на разхода (условно)."

    elif intent == "FISCAL_INVEST":
        exp_df.loc[exp_df["Category"] == "Capex (public investment)", "Amount (bn BGN)"] += 1.0
        exp_df.loc[exp_df["Category"].isin(["Education", "Healthcare"]), "Amount (bn BGN)"] += 0.3
        note = "DEMO сценарий: инвестиции → +1.0 млрд капекс и +0.3 млрд образование/здраве (условно)."

    total_rev = float(rev_df["Amount (bn BGN)"].sum())
    total_exp = float(exp_df["Amount (bn BGN)"].sum())
    deficit = total_exp - total_rev

    deficit_pct = (deficit / gdp) if gdp else None
    debt_pct = (debt / gdp) if (gdp and debt is not None) else None

    st.subheader("Финансов резултат (DEMO)")
    a, b, c, d = st.columns(4)
    a.metric("Приходи", f"{total_rev:.1f} млрд. лв.")
    b.metric("Разходи", f"{total_exp:.1f} млрд. лв.")
    c.metric("Дефицит", f"{deficit:.1f} млрд. лв.")
    d.metric("Дефицит (% БВП)", f"{deficit_pct*100:.2f}%" if deficit_pct is not None else "n/a")

    f_light, d_light = traffic(deficit_pct, debt_pct, goal_def=goal_def, goal_debt=goal_debt)
    st.write(f"Цели: дефицит ≤ 3% и дълг ≤ 60% → Светофар: **Дефицит {f_light} | Дълг {d_light}**")
    st.info(note)

    gap = max(aic_eu - aic_bg, 0) if (aic_eu is not None and aic_bg is not None) else None
    st.caption(
        f"AIC (DEMO): BG={aic_bg:.1f}, EU={aic_eu:.1f}, gap={gap:.1f} пункта"
        if gap is not None else "AIC (DEMO): n/a"
    )

    st.divider()
    left, right = st.columns(2)
    with left:
        st.subheader("Приходи (след сценария)")
        st.dataframe(rev_df, use_container_width=True, hide_index=True)
    with right:
        st.subheader("Разходи (след сценария)")
        st.dataframe(exp_df, use_container_width=True, hide_index=True)

    # ---- AI analysis (real-time) using computed KPIs + user question
    system = f"""
Ти си BGGovAI — аналитичен съветник за публични политики на България.

{GOALS_TEXT}

Правила:
- Отговаряй кратко, структурирано и прагматично.
- Ползвай числата от модела (дефицит/дълг/AIC gap).
- Покажи trade-offs и как се спазват целите.
- Не измисляй данни, които не са дадени.
"""

    context = f"""
Въпрос от потребителя:
{q}

Ключови индикатори (от Excel демо модела):
- Приходи: {total_rev:.1f} млрд. лв.
- Разходи: {total_exp:.1f} млрд. лв.
- Дефицит: {deficit:.1f} млрд. лв.
- Дефицит (% БВП): {deficit_pct*100:.2f}%  (цел ≤ 3%)
- Дълг (% БВП): {(debt_pct*100):.2f}%  (цел ≤ 60%)  (ако има данни)
- AIC BG: {aic_bg:.1f} / AIC EU: {aic_eu:.1f} / Gap: {gap:.1f}

Политическо ограничение (демо): без повишаване на данъците.
"""

    st.divider()
    st.subheader("AI анализ (real-time)")
    st.write(ask_ai(system, context))
    st.caption("Ако искаш друг модел: в Secrets добави OPENAI_MODEL=\"...\" (пример: gpt-4o-mini).")


elif intent == "ADMIN_MOL":
    answer_admin_mol()
    st.divider()
    st.subheader("AI допълнение (по желание, real-time)")
    system = f"Ти си административен консултант. {GOALS_TEXT}\nОтговаряй ясно и по стъпки."
    context = f"Въпрос: {q}\nДай практичен чеклист и какви документи трябват."
    st.write(ask_ai(system, context))


elif intent == "LEGAL_CITIZENSHIP":
    answer_legal_citizenship()
    st.divider()
    st.subheader("AI допълнение (по желание, real-time)")
    system = "Ти си правен анализатор. Отговаряй структурирано, без да измисляш конкретни членове."
    context = f"Въпрос: {q}\nДай рамка, рискове, и какви данни/текст липсват за точен анализ."
    st.write(ask_ai(system, context))


else:
    # GENERAL: real-time AI chat with guardrails + mention supported topics
    st.subheader("Общ отговор (real-time AI)")
    system = f"""
Ти си BGGovAI — ИИ съветник за публични политики (демо).
{GOALS_TEXT}

Ограничения:
- Ако въпросът е фискален и няма Excel — кажи, че липсва бюджет.
- Ако темата е извън демото — предложи какви данни трябват.
- Не измисляй факти.
"""
    context = f"""
Въпрос: {q}

Поддържани теми в демото (за ориентация):
- {", ".join(SUPPORTED)}
"""
    st.write(ask_ai(system, context))
