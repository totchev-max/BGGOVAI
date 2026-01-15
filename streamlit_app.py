
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from pathlib import Path
import base64

st.set_page_config(page_title="Република България — BGGovAI (DEMO)", layout="wide")

ASSETS = Path(__file__).parent / "assets"

def b64(path: Path) -> str:
    return base64.b64encode(path.read_bytes()).decode("utf-8")

CREST_B64 = b64(ASSETS / "crest_demo.svg")

st.markdown(
    f"""
    <style>
      .gov-header {{
        border-radius: 14px;
        overflow: hidden;
        box-shadow: 0 2px 10px rgba(0,0,0,0.08);
        border: 1px solid rgba(0,0,0,0.06);
        margin-bottom: 14px;
      }}
      .flag {{
        height: 10px;
        background: linear-gradient(to bottom, #ffffff 0%, #ffffff 33%, #00966E 33%, #00966E 66%, #D62612 66%, #D62612 100%);
      }}
      .gov-top {{
        display: flex;
        gap: 14px;
        align-items: center;
        padding: 14px 16px;
        background: #ffffff;
      }}
      .crest {{
        width: 54px;
        height: 54px;
      }}
      .gov-title {{
        line-height: 1.1;
      }}
      .gov-title h1 {{
        margin: 0;
        font-size: 18px;
        font-weight: 700;
      }}
      .gov-title p {{
        margin: 4px 0 0 0;
        font-size: 13px;
        color: rgba(0,0,0,0.65);
      }}
      .disclaimer {{
        border-radius: 12px;
        padding: 10px 12px;
        background: rgba(214,38,18,0.06);
        border: 1px solid rgba(214,38,18,0.20);
        font-size: 13px;
        margin-bottom: 10px;
      }}
      .chip {{
        display:inline-block;
        padding: 3px 8px;
        border-radius: 999px;
        background: rgba(0,150,110,0.08);
        border: 1px solid rgba(0,150,110,0.22);
        font-size: 12px;
        margin-right: 6px;
        margin-bottom: 6px;
      }}
    </style>

    <div class="gov-header">
      <div class="flag"></div>
      <div class="gov-top">
        <img class="crest" src="data:image/svg+xml;base64,{CREST_B64}" />
        <div class="gov-title">
          <h1>Република България — BGGovAI</h1>
          <p>ИИ съветник за публични политики (демонстрационна версия)</p>
        </div>
      </div>
    </div>

    <div class="disclaimer">
      <b>Внимание:</b> Това е <b>демо прототип</b>. Не е официален държавен портал и не представлява правен/финансов съвет.
      „Гербът“ тук е <b>стилизиран демо символ</b>. При нужда може да се замени с официално изображение при разрешение.
    </div>
    """,
    unsafe_allow_html=True,
)

uploaded = st.file_uploader("Качи Excel бюджет (.xlsx)", type=["xlsx"])

SUPPORTED = [
    "ДДС 9% за ресторанти (въздействие)",
    "Пенсии +10% (въздействие)",
    "Инвестиции (Capex+образование+здраве)",
    "Закон за българското гражданство (рамка за анализ)",
    "Смяна на МОЛ на ЕООД (документи и стъпки)",
    "Общ фискален преглед (дефицит/дълг/AIC)",
]

def classify(q: str) -> str:
    t = q.lower()
    if any(k in t for k in ["мол", "управител", "еоод", "търговски регист", "а4"]):
        return "ADMIN_MOL"
    if any(k in t for k in ["гражданств", "закон за българското гражданство", "натурализ", "изменени", "проект"]):
        return "LEGAL_CITIZENSHIP"
    if "ддс" in t and any(k in t for k in ["ресторан", "кетър", "9%"]):
        return "FISCAL_VAT_REST"
    if "пенс" in t and any(k in t for k in ["10", "процент", "%"]):
        return "FISCAL_PENSIONS"
    if any(k in t for k in ["инвест", "капекс", "инфраструкт", "образован", "здравеопаз"]):
        return "FISCAL_INVEST"
    if any(k in t for k in ["дефиц", "дълг", "бюджет", "бвп", "aic", "догон"]):
        return "FISCAL_BASE"
    return "UNKNOWN"

def answer_admin_mol():
    st.subheader("Администрация: Смяна на МОЛ (управител) на ЕООД — DEMO чеклист")
    st.markdown("""
**Къде:** Търговски регистър (Агенция по вписванията)  
**Заявление:** обичайно **А4** (промени по обстоятелства)  
**Документи (типично):**
- Решение на едноличния собственик за освобождаване/назначаване на управител
- Съгласие и образец от подпис (спесимен) на новия управител (често с нотариална заверка)
- Декларации по ТЗ (според конкретиката)
- Такса за вписване (електронно е по-евтино)
**Стъпки:**
1) Подготовка на решение/декларации/подпис
2) Подаване в ТР (КЕП или на място)
3) Проверка на вписването + уведомяване на банки/партньори
""")
    st.caption("Бележка: демо ориентир. Реалният пакет документи зависи от казуса и изискванията към заверките.")

def answer_legal_citizenship():
    st.subheader("Право: Закон за българското гражданство — DEMO рамка за анализ")
    st.markdown("""
**Структура за оценка на предложение:**
1) Какво точно се променя (условия, срокове, изключения)  
2) Съответствие с Конституция/международни ангажименти  
3) Процедури и административна изпълнимост (капацитет, срокове, контрол)  
4) Рискове: неясни дефиниции, обжалвания, конфликт на норми, преходни режими  
5) Мерки за минимизиране: ясни дефиниции, преходни разпоредби, подзаконови актове, ИТ промени
""")
    st.caption("За конкретика: нужен е текстът на предложението (чл./ал./§), за да се маркират точните изменения.")

def table_to_df(rows, total_keyword="TOTAL"):
    header = None
    body = []
    for r in rows:
        if r and r[0] == "Category" and r[1] == "Amount (bn BGN)":
            header = list(r[:3])
            continue
        if header and r and r[0]:
            body.append(list(r[:3]))
    df = pd.DataFrame(body, columns=header or ["Category","Amount (bn BGN)","Notes"])
    df = df[~df["Category"].astype(str).str.contains(total_keyword, na=False)]
    df["Amount (bn BGN)"] = pd.to_numeric(df["Amount (bn BGN)"], errors="coerce").fillna(0.0)
    return df

def parse_inputs(rows):
    vals = {}
    for r in rows:
        if not r or not r[0]:
            continue
        vals[str(r[0]).strip()] = r[1]
    def getf(k, default=None):
        v = vals.get(k, default)
        try:
            return float(v)
        except Exception:
            return default
    return {
        "gdp": getf("GDP (bn BGN)", None),
        "debt": getf("Debt stock (bn BGN)", None),
        "aic_bg": getf("AIC (EU=100) - Bulgaria", 70.0),
        "aic_eu": getf("AIC (EU=100) - EU average", 100.0),
    }

def traffic(deficit_pct, debt_pct, goal_def=0.03, goal_debt=0.60):
    def light(val, green_th, yellow_th):
        if val is None:
            return "⚪️"
        if val <= green_th:
            return "🟩"
        if val <= yellow_th:
            return "🟨"
        return "🟥"
    f = light(abs(deficit_pct) if deficit_pct is not None else None, goal_def, goal_def*1.5)
    d = light(debt_pct, goal_debt, goal_debt+0.10)
    return f, d

st.markdown("### Въпрос към системата")
q = st.text_area("Пиши свободно (демото разпознава валидиран набор теми)", height=90)

st.markdown(" ".join([f'<span class="chip">{s}</span>' for s in SUPPORTED]), unsafe_allow_html=True)
do = st.button("Отговори", use_container_width=True)

if not do:
    st.stop()

intent = classify(q)

if intent.startswith("FISCAL") and not uploaded:
    st.warning("За финансовите въпроси първо качи Excel бюджета.")
    st.stop()

if intent.startswith("FISCAL"):
    wb = load_workbook(filename=BytesIO(uploaded.getvalue()), data_only=True)
    need = {"Inputs","Revenues","Expenditures"}
    if not need.issubset(set(wb.sheetnames)):
        st.error("Липсват листове. Нужни: Inputs, Revenues, Expenditures.")
        st.stop()

    inp = parse_inputs(list(wb["Inputs"].values))
    gdp = inp["gdp"]
    debt = inp["debt"]
    aic_bg = inp["aic_bg"]
    aic_eu = inp["aic_eu"]

    rev_df = table_to_df(list(wb["Revenues"].values), total_keyword="TOTAL")
    exp_df = table_to_df(list(wb["Expenditures"].values), total_keyword="TOTAL")

    goal_def = 0.03
    goal_debt = 0.60
    note = "DEMO: общ фискален преглед."

    if intent == "FISCAL_VAT_REST":
        rev_df.loc[rev_df["Category"]=="VAT (total)", "Amount (bn BGN)"] -= 0.6
        note = "DEMO сценарий: ДДС 9% за ресторанти → -0.6 млрд. лв. приход от ДДС (условно)."
    elif intent == "FISCAL_PENSIONS":
        exp_df.loc[exp_df["Category"]=="Pensions", "Amount (bn BGN)"] *= 1.10
        note = "DEMO сценарий: +10% пенсии (условно увеличение на разхода)."
    elif intent == "FISCAL_INVEST":
        exp_df.loc[exp_df["Category"]=="Capex (public investment)", "Amount (bn BGN)"] += 1.0
        exp_df.loc[exp_df["Category"].isin(["Education","Healthcare"]), "Amount (bn BGN)"] += 0.3
        note = "DEMO сценарий: инвестиции → +1.0 млрд капекс и +0.3 млрд образование/здраве (условно)."

    total_rev = float(rev_df["Amount (bn BGN)"].sum())
    total_exp = float(exp_df["Amount (bn BGN)"].sum())
    deficit = total_exp - total_rev

    deficit_pct = deficit / gdp if gdp else None
    debt_pct = debt / gdp if (gdp and debt is not None) else None

    st.subheader("Финансов резултат (DEMO)")
    a,b,c,d = st.columns(4)
    a.metric("Приходи", f"{total_rev:.1f} млрд. лв.")
    b.metric("Разходи", f"{total_exp:.1f} млрд. лв.")
    c.metric("Дефицит", f"{deficit:.1f} млрд. лв.")
    d.metric("Дефицит (% БВП)", f"{deficit_pct*100:.2f}%" if deficit_pct is not None else "n/a")

    f_light, d_light = traffic(deficit_pct, debt_pct, goal_def=goal_def, goal_debt=goal_debt)
    st.write(f"Цели: дефицит ≤ 3% и дълг ≤ 60% → Светофар: Дефицит {f_light} | Дълг {d_light}")
    st.info(note)

    st.caption(f"AIC (DEMO): BG={aic_bg:.1f}, EU={aic_eu:.1f}, gap={max(aic_eu-aic_bg,0):.1f} пункта")

    st.divider()
    left, right = st.columns(2)
    with left:
        st.subheader("Приходи (след сценария)")
        st.dataframe(rev_df, use_container_width=True, hide_index=True)
    with right:
        st.subheader("Разходи (след сценария)")
        st.dataframe(exp_df, use_container_width=True, hide_index=True)

elif intent == "ADMIN_MOL":
    answer_admin_mol()
elif intent == "LEGAL_CITIZENSHIP":
    answer_legal_citizenship()
else:
    st.warning("Демо прототипът разпознава ограничен набор теми. Поддържани въпроси/теми:")
    st.write("• " + "\n• ".join(SUPPORTED))
